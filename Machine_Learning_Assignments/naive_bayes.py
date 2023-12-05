# -*- coding: utf-8 -*-
"""
Created on Sat Nov  5 11:26:59 2022

@author: ozmen
"""
from collections import defaultdict
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from math import pi, exp

def mean_sd(attr):
    summ=0.0
    for value in attr:
        summ += value
    mean = summ / len(attr)
    summ=0.0
    for value in attr:
        summ+=(value - mean)**2
    sd = (summ / len(attr))**0.5
    return mean, sd

def probability(value, mean, sd):
    sigmaroottwo = sd * (2**0.5)
    return (1/(sigmaroottwo*(pi**0.5))) * exp(-((value-mean)/sigmaroottwo)**2)

def get_categories(y_train):
    cat_len = len(y_train)
    categories = defaultdict(lambda : 0)
    for y in y_train:
        categories[y] += 1
    categories = {k: v / cat_len for k, v in categories.items()}
    return categories    
    
def attr_category_table (x_train, y_train):
    categories = get_categories(y_train)
    attr_category_table = defaultdict(lambda : 0)
    attributes = [[x_train[i][j] for i in range(len(x_train))] for j in \
                  range(len(x_train[0]))] 
    for index, attr in enumerate(attributes):
        for category in categories:
            subset=[value for i, value in enumerate(attr) if\
                    y_train[i] == category]
            attr_category_table[index, category] = mean_sd(subset)
    return attr_category_table       

def get_predictions(table, x_test, categories):
    predictions = list()
    for record in x_test:
        probabilities = dict()
        for category in categories:
            prob = categories[category]
            for i, value in enumerate(record):
                mean, sd = table[i, category]
                prob *= probability(value, mean, sd)
            probabilities[category] = prob
        predictions.append(max(probabilities, key = probabilities.get)) 
    return predictions   

def normalize(raw_set, norm_set, epsilon = 0.000001):
    """
    Parameters
    ----------
    raw_set : TYPE
        Set to be normalized.
    norm_set : TYPE
        Normalizing set. Is used for choosing min, max values
        epsilon : TYPE, optional
        To 'remove possibility'/'reduce likelihood' of division by zero. 
        The default is 0.000001.
    Returns normalized_set 
    """
    norm_set_minmax=[(min(att), max(att)) for att in [[att_set[i] for att_set
                                                       in norm_set] 
                                                      for i in 
                                                      range(len(norm_set[0]))]]
    normalized_set = list()
    for raw_att in raw_set:
        normalized_att = list()
        for att, norm_att_minmax in zip(raw_att, norm_set_minmax):
            normalized_att.append((att-norm_att_minmax[0])/(norm_att_minmax[1]
                                                            -norm_att_minmax[0]
                                                            +epsilon))
        normalized_set.append(normalized_att)
    return normalized_set   

def get_accuracy(predictions, y_test):
    accuracy_dict = defaultdict(lambda : 0) 
    for prediction, target in zip(predictions, y_test):
        if prediction == target:
            accuracy_dict['true'] += 1
        else:
            accuracy_dict['false'] += 1
    return accuracy_dict['true']/(accuracy_dict['true']+accuracy_dict['false'])

        
x_data, y_data = (list(), list())
wb = load_workbook('Dry_Bean_Dataset.xlsx')
ws = wb["Dry_Beans_Dataset"]
for row in ws.iter_rows(min_row = 2, values_only = True):
    x_data.append(list())
    for i in range(16):
        x_data[-1].append(row[i])
    y_data.append(row[16])

x_train, x_test, y_train, y_test = train_test_split(x_data, y_data,
                                                            test_size=0.3,
                                                            random_state=0)

x_test = normalize(x_test, x_train)
x_train = normalize(x_train, x_train)

table = attr_category_table(x_train, y_train)
categories = get_categories(y_train)
predictions = get_predictions(table, x_test, categories)
accuracy = get_accuracy(predictions, y_test)

print(f"Test set cccuracy = {accuracy:.4%}")








        