# -*- coding: utf-8 -*-
"""
Created on Thu Nov  3 17:07:20 2022

@author: ozmen
"""
from collections import defaultdict
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from functools import partial

DATASET = r".\Dry_Bean_Dataset.xlsx"

def normln(x, y, n = 2):
    """
    Returns norm L'n' of x-y
    """
    summary = 0.0
    for i in range(len(x)):
        summary += abs(x[i] - y[i])**n
    return summary**(1/n)

def get_neighbours(x_train, y_train, x_test, k, distance_function = normln):
    """
    Returns distances of k nearest neighbours and their labels
    """
    neighbours = list()
    for point in x_test:
        point_neighbours = list()
        for x, y in zip(x_train, y_train):
            point_neighbours.append((distance_function(x, point), y))
        point_neighbours.sort(key = lambda tupl: tupl[0])
        neighbours.append(point_neighbours[:k])
    return neighbours    

def predict(point_neighbours, epsilon):
    """
    Returns predicted label, given a neighbour distance and label list
    """
    weighted_sums = defaultdict(lambda : 0)
    for neighbour in point_neighbours:
        weighted_sums[neighbour[1]] += 1 / (neighbour[0] + epsilon)**2
    return max(weighted_sums, key = weighted_sums.get)

def get_predictions(neighbours, epsilon = 0.000001):
    """
    Returns predictions, given neighbours of a set
    """
    return [predict(neighbours[i], epsilon) for i in range(len(neighbours))]
    
def get_accuracy(predictions, y_test):
    accuracy_dict = defaultdict(lambda : 0) 
    for prediction, target in zip(predictions, y_test):
        if prediction == target:
            accuracy_dict['true'] += 1
        else:
            accuracy_dict['false'] += 1
    return accuracy_dict['true']/(accuracy_dict['true']+accuracy_dict['false'])

def get_k(x_train, y_train, x_val, y_val, min_k = 1, max_k = 50,
          distance_function = normln):
    neighbours_max_k = get_neighbours(x_train, y_train, x_val, max_k,
                                      distance_function)
    accuracy_list = list()
    for k in range(min_k, max_k + 1):
        accuracy=get_accuracy(get_predictions([neighbours_max_k[i][:k] for 
                                               i in 
                                               range(len(neighbours_max_k))]),
                              y_val)
        accuracy_list.append(accuracy)
    return (min_k + max(range(len(accuracy_list)), 
                       key=accuracy_list.__getitem__),
           max(accuracy_list))

def normalize(raw_set, norm_set, epsilon = 0.000001):
    """
    Parameters
    ----------
    raw_set : TYPE
        Set to be normalized.
    norm_set : TYPE
        Normalizing set. Is used for choosing min, max values
        epsilon : TYPE, optional
        To 'remove possibility'/'reduce likelihood' of division by zero. 
        The default is 0.000001.
    Returns normalized_set 
    """
    norm_set_minmax=[(min(att), max(att)) for att in [[att_set[i] for att_set
                                                       in norm_set] 
                                                      for i in 
                                                      range(len(norm_set[0]))]]
    normalized_set = list()
    for raw_att in raw_set:
        normalized_att = list()
        for att, norm_att_minmax in zip(raw_att, norm_set_minmax):
            normalized_att.append((att-norm_att_minmax[0])/(norm_att_minmax[1]
                                                            -norm_att_minmax[0]
                                                            +epsilon))
        normalized_set.append(normalized_att)
    return normalized_set    

x_data, y_data = (list(), list())
wb = load_workbook('Dry_Bean_Dataset.xlsx')
ws = wb["Dry_Beans_Dataset"]
for row in ws.iter_rows(min_row = 2, values_only = True):
    x_data.append(list())
    for i in range(16):
        x_data[-1].append(row[i])
    y_data.append(row[16])

x_train_val, x_test, y_train_val, y_test = train_test_split(x_data, y_data,
                                                            test_size=0.3,
                                                            random_state=0)
x_train, x_val, y_train, y_val = train_test_split(x_train_val, y_train_val,
                                                  test_size=0.2,
                                                  random_state=0)

x_test = normalize(x_test, x_train)
x_val = normalize(x_val, x_train)
x_train = normalize(x_train, x_train)

accuracy = 0
for norm in range(1,5):
    k_new, accuracy_new = get_k(x_train, y_train, x_val, y_val, 1, 200,
                        partial(normln, n = norm))
    print(f"Best accuracy for norm = {norm} for the validation set is with"
          f" k = {k_new}. Validation accuracy = {accuracy_new:.4%}") 
    if(accuracy_new > accuracy):
        k, n = k_new, norm
        accuracy = accuracy_new

accuracy = get_accuracy(get_predictions(
    get_neighbours(x_train, y_train, x_test, k, partial(normln, n = n))), 
    y_test)

print(f"Accuracy for the test set on KNN classifier using k = {k}"
      f" with norm = {n} is {accuracy:.4%}")        
        
    









    
    
    